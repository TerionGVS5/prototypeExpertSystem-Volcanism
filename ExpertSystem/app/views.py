"""
Definition of views.
"""

from django.shortcuts import render
from app.forms import *
from django.http import HttpRequest
from django.template import RequestContext
from datetime import datetime
from django.views import View
from django.http import HttpResponseRedirect
from app.models import *
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.views.generic.edit import UpdateView
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist
import os
from django.http import HttpResponse
from django.http import JsonResponse
import copy

def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/index.html',
        {
            'title':'Home Page',
            'year':datetime.now().year,
        }
    )

def maps(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    volcanoesFullInfo = Volcano.objects.all().values_list('pk','name', 'latitude', 'longitude', 'activ')
    volcanoesJsonFullInfo = json.dumps(list(volcanoesFullInfo), cls=DjangoJSONEncoder)
    return render(
        request,
        'app/maps.html',
        {
            'title':'Maps',
            'volcanoesJsonFullInfo':volcanoesJsonFullInfo,
            'message':'Your contact page.',
            'year':datetime.now().year,
        }
    )

def signs(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)

    return render(
        request,
        'app/signs.html',
        {
            'title':'Signs',
            'message':'Your contact page.',
            'year':datetime.now().year,
        }
    )

def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/contact.html',
        {
            'title':'Contact',
            'message':'Your contact page.',
            'year':datetime.now().year,
        }
    )

def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'app/about.html',
        {
            'title':'About',
            'message':'Your application description page.',
            'year':datetime.now().year,
        }
    )

class Addsign(View):
    classform = SignAddForm
    template_name='app/addsign.html'
    title='Добавление Свойств'
    def get(self,request):
        form = self.classform()
        return render(request,self.template_name,{'title':self.title,'form':form})
    def post(self,request):
        form = self.classform(request.POST)
        new_sign = form.save()
        return HttpResponseRedirect('/addsign')

class Addvalue(View):
    classform = ValueAddForm
    template_name='app/addvalue.html'
    title='Установка значений свойств'
    def get(self,request):
        form = self.classform()
        return render(request,self.template_name,{'title':self.title,'form':form})
    def post(self,request):
        form = self.classform(request.POST)
        new_sign = form.save()
        return HttpResponseRedirect('/addvalue')

class Addgroupvolcano(View):
    classform = GroupVolcanoAddForm
    template_name='app/addgroupvolcano.html'
    title='Добавление груп вулканов'
    def get(self,request):
        form = self.classform()
        return render(request,self.template_name,{'title':self.title,'form':form})
    def post(self,request):
        form = self.classform(request.POST)
        new_sign = form.save()
        return HttpResponseRedirect('/addgroupvolcano')

class Addvolcano(View):
    classform = VolcanoAddForm
    template_name='app/addvolcano.html'
    title='Добавление вулканов'
    def get(self,request):
        form = self.classform()
        return render(request,self.template_name,{'title':self.title,'form':form})
    def post(self,request):
        form = self.classform(request.POST)
        new_sign = form.save()
        return HttpResponseRedirect('/addvolcano')

class VolcanoUpdate(UpdateView):
    model = Volcano
    form_class = VolcanoUpdateForm
    

    template_name_suffix = '_update_form'
    success_url = '/'

class SelectVolcano(View):
    classform = VolcanoSelectForm
    template_name='app/selectvolcano.html'
    title='Выбор вулканов'
    def get(self,request):
        form = self.classform()
        return render(request,self.template_name,{'title':self.title,'form':form})
    def post(self,request):
        form = self.classform(request.POST)
        if form.is_valid():
            key = form.cleaned_data['volcano'].pk
            return HttpResponseRedirect('/updatevolcano/{}/'.format(key))

def fillDB(request):
    data_values = Value.objects.all().count()
    if data_values !=0:
        return HttpResponse("В базе уже что-то есть. Нужно очистить сначала")
    else:
        path = os.path.abspath("data.xlsx")
        wb = load_workbook(path)
        ws = wb.active
        row_num = 0
        for row in ws.iter_rows():
            row_num+=1
            column_num=0
            for cell in row:
                column_num+=1
                try:
                    volcan_now = Volcano.objects.get(pk=row_num)
                    sign_now = Sign.objects.get(pk=column_num)
                except ObjectDoesNotExist:
                    continue
                value_now = Value(value = cell.value,volcano = volcan_now, sign = sign_now)
                value_now.save()

        return HttpResponse("Вроде заполнилось всё")

def getInfoVolcano(request):
    key = request.GET.get('key')
    volcano_info = Volcano.objects.get(pk=key)
    data = {}
    data['name']=volcano_info.name
    data['latitude']=volcano_info.latitude
    data['longitude']=volcano_info.longitude
    data['activ']=volcano_info.activ
    data['description']=volcano_info.description
    data['image']=str(volcano_info.image)
    json_data=json.dumps(data, cls=DjangoJSONEncoder)
    return HttpResponse(json_data, content_type='application/json')

def onegraph(request):
    l_count = int(request.GET.get('l_count'))
    array_volcano_id = json.loads(request.GET.get('array_volcano_id'))
    array_sign_id = json.loads(request.GET.get('array_sign_id'))
    matrix_h = [[0 for j in range(0,len(array_sign_id))] for i in range(0,len(array_volcano_id))]
    who_is_who = {}
    for volcano_id in array_volcano_id:
        for sign_id in array_sign_id:
            matrix_h[array_volcano_id.index(volcano_id)][array_sign_id.index(sign_id)] = int(Value.objects.get(volcano = volcano_id, sign = sign_id).value)
            who_is_who[array_volcano_id.index(volcano_id)] = volcano_id

    def get_diff(string1, string2):
        diff = 0
        for index in range(0, len(string1)):
            if string1[index] != string2[index]:
                diff += 1
        return diff


    def make_new_matrix(input_matrix):
        new_matrix = []
        for string in input_matrix:
            string_to_append = []
            for next_s in input_matrix:
                diff = get_diff(string, next_s)
                string_to_append.append(diff)
            new_matrix.append(string_to_append)
        return new_matrix


    def remove_diagonal(input_matrix):
        for i in range(0, len(input_matrix)):
            for j in range(0, len(input_matrix)):
                if i == j:
                    input_matrix[i][j] = -1
        return input_matrix


    def remove_index(input_matrix, index_to_remove):
        for i in range(0, len(input_matrix)):
            for j in range(0, len(input_matrix)):
                if i == index_to_remove or j == index_to_remove:
                    input_matrix[i][j] = -1
        return input_matrix


    def first_find_min(input_matrix):
        min = 1000
        way = []
        distances = []
        index_to_remove = 0
        index_to_continue = 0
        for i in range(0, len(input_matrix)):
            for j in range(0, len(input_matrix)):
                if input_matrix[i][j] == -1:
                    continue
                if input_matrix[i][j] < min:
                    min = input_matrix[i][j]
                    index_to_remove = i
                    index_to_continue = j
        new_matrix = remove_index(input_matrix, index_to_remove)
        way.append(index_to_remove + 1)
        distances.append(min)
        return new_matrix, index_to_continue, way, distances


    def append_last_way(way, input_matrix):
        for index in range(0, len(input_matrix)):
            if index + 1 not in way:
                new_way = way
                new_way.append(index + 1)
                return new_way


    def just_find_min(input_matrix, index_to_continue, way, distances):
        min = 1000
        index_to_remove = 0
        new_index_to_continue = 0
        for j in range(0, len(input_matrix)):
            if input_matrix[index_to_continue][j] == -1:
                continue
            if input_matrix[index_to_continue][j] < min:
                min = input_matrix[index_to_continue][j]
                index_to_remove = index_to_continue
                new_index_to_continue = j
        if min == 1000:
            new_way = append_last_way(way, input_matrix)
            return new_way, distances
        else:
            new_matrix = remove_index(input_matrix, index_to_remove)
            new_way = way
            new_way.append(index_to_remove + 1)
            new_distances = distances
            new_distances.append(min)
            return just_find_min(new_matrix, new_index_to_continue, new_way, new_distances)


    def make_sets(way, distances, l_count):
        r = l_count - 1
        exclude = []
        result_sets = []
        while len(exclude) != r:
            max = -1
            index_to_exclude = -1
            for index in range(0, len(distances)):
                if index in exclude:
                    continue
                if distances[index] > max:
                    max = distances[index]
                    index_to_exclude = index
            exclude.append(index_to_exclude)
        set_to_append = []
        for index in range(0, len(way)):
            set_to_append.append(way[index])
            if index in exclude:
                result_sets.append(set_to_append)
                set_to_append = []
        result_sets.append(set_to_append)
        return result_sets


    def main(MATRIX_H, L_COUNT):
        matrix = make_new_matrix(MATRIX_H)
        obj_matrix = copy.deepcopy(matrix)
        matrix = remove_diagonal(matrix)
        matrix, index_to_continue, way, distances = first_find_min(matrix)
        way, distances = just_find_min(matrix, index_to_continue, way, distances)
        sets=make_sets(way, distances, L_COUNT)
        return obj_matrix,way,distances,sets
    obj_matrix,way,distances,sets = main(matrix_h,l_count)
    new_sets = copy.deepcopy(sets)
    for i in range(0,len(sets)):
        for j in range(0,len(sets[i])):
            new_sets[i][j] = who_is_who[sets[i][j]-1]
    json_data=json.dumps(new_sets, cls=DjangoJSONEncoder)
    return HttpResponse(json_data, content_type='application/json')

def getInfoSign(request):
    '''
    SignFullInfo = Sign.objects.all().values_list('name', 'number', 'groupsign')
    ListSignFullInfo = list(SignFullInfo)
    ListSignFullInfo = [list(el) for el in ListSignFullInfo]
    for element in ListSignFullInfo:
        element[2] = GroupSign.objects.get(pk=element[2]).name
    SignJsonFullInfo = json.dumps(ListSignFullInfo, cls=DjangoJSONEncoder)
    return HttpResponse(SignJsonFullInfo, content_type='application/json')
    '''
    groups = GroupSign.objects.all()
    data = {}
    for group in groups:
        signs = Sign.objects.filter(groupsign=group)
        data[group.name]=[]
        for sign in signs:
            data[group.name].append({sign.name:sign.number})
    data_json = json.dumps(data, cls=DjangoJSONEncoder)
    return HttpResponse(data_json, content_type='application/json')

def masks(request):
    array_volcano_id = json.loads(request.GET.get('array_volcano_id'))
    array_sign_id = json.loads(request.GET.get('array_sign_id'))

    T = [[int(i)] for i in array_volcano_id]

    for sign_id in array_sign_id:
        for el in T:
            el.append(int(Value.objects.get(volcano = el[0], sign = sign_id).value))
    print(T)

    def MaskMethod(T) :
	    ##T - матриа признаков, которая поступает на вход, первый столбец матрицы - это айдишники вулканов из БД 
	    #T = [  #ID X1  X2  X3  X4  X5  X6  X7  X8  X9  X10 X11 X12
	    #	 [ 1,  1,  1,  1,  1], #A1
	    #	 [ 2,  0,  0,  0,  0], #A2
	    #	 [ 3,  1,  0,  1,  0], #A3
	    #	 [ 4,  1,  1,  1,  1], #A4 
	    #	 ]

        m = 0 #Количество строк
        for i in T :
            m += 1 

        n = 0 #Количество колонок 
        for i in T[0] :
            n += 1
        n -= 1

        Table = [["null"] * (n+1) for i in range(m+1)]  

        for i in range(1,n+1) :
            Table[0][i] = "X"+str(i)

        for i in range(1,m+1) :
            Table[i][0] = "A"+str(i)

        for i in range(1, m+1) :
            for j in range(1, n+1) :
                Table[i][j] = T[i-1][j]

        R1 = round((n / 2) + (n / 4)) #Порог различимости 
        R2 = round((n / 2) + (n / 4)) #Порог сходства 

        MAT = [[0] * (n+1) for i in range(m)]
        MATprev = []
        for i in range(n) :
            MATprev.append(0)
        MATtmp = []
        for i in range(n) :
            MATtmp.append(0)
        differenceR1 = 0
        differenceR2 = 0 

        for i in range(n) : 
            MAT[0][i]=T[0][i+1]
            MATprev[i]=T[0][i+1]

        MAT[0][n]=1

        for i in range(1,m):
            for k in range(m-1):
                for g in range (n):
                    if MAT[k][g] != T[i][g+1] :
                        differenceR2+=1;
                if differenceR2 < R2 :
                    break
            for j in range(n) :
                if T[i][j+1] != MATprev[j] :
                    differenceR1+=1
                MATtmp[j]=T[i][j+1]
            if differenceR1 >= R1 and differenceR2 >= R2 : 
                for j in range(n) :
                    MAT[i][j]= MATtmp[j]
                    MATprev[j]= MATtmp[j]
                MAT[i][n] = 1
            MATtmp = []
            for i in range(n) :
                MATtmp.append(0)
            differenceR2=0
            differenceR1=0

        for elem in Table :
            elem.append("null")
        len(Table[0])
        Table[0][len(Table[0])-1] = 'M'


        j = 1
        for i in range(1,m):
            if (MAT[i-1][n] == 1) :
                Table[i][len(Table[0])-1] = 'M'+str(j)
                j += 1

        j = 1
        for i in range(m-1) :
            if (MAT[i][n] == 1) :
                if (j==1) :
                    codesBegin = len(Table[0])
                for elem in Table :
                    elem.append("null")
                Table[0][len(Table[0])-1] = 'M'+ str(j)
                j += 1
        codesEnd = len(Table[0])-1 

        k = 0
        for i in range(m-1) :
            if MAT[i][n] == 1 :
                for g in range (m) :
                    for j in range (n) :
                        if T[g][j+1] == MAT[i][j] : 
                            differenceR2 += 1
                    if differenceR2 >= R2 :
                        Table[g+1][codesBegin+k] = '1'
                    else :
                        Table[g+1][codesBegin+k] = '0'
                    differenceR2 = 0
                k += 1

        i = 0
        Clusters = [["null"] * (2) for i in range(m+1)]
        tmpCluster = ""
        for i in range(1,m+1) :
            for j in range(codesBegin, codesEnd+1) :
                tmpCluster = tmpCluster + Table[i][j]
            Clusters[i-1][0] =tmpCluster #Значение
            Clusters[i-1][1] = i         #Номер
            tmpCluster = ""

        tmpCluster1 = []
        for i in range(2) :
            tmpCluster1.append("null")
        for j in range(m-1) :
            for i in range(m-j-1) : 
                if (Clusters[i][0] > Clusters[i+1][0]) : 
                    tmpCluster1[0] = Clusters[i][0]
                    tmpCluster1[1] = Clusters[i][1]
                    Clusters[i][0] = Clusters[i+1][0]
                    Clusters[i][1] = Clusters[i+1][1]
                    Clusters[i+1][0] =  tmpCluster1[0]
                    Clusters[i+1][1] =  tmpCluster1[1]


        l =0
        Res = ""
        for i in range(m+1) :
            if (Clusters[i-1][0] != Clusters[i][0]) and Clusters[i][0] != "null" :
                if i>=1 :
                    Res = Res + " "
                l += 1
            if Clusters[i][1] != "null" :
                Res =  Res + 'a' + str(Clusters[i][1])
        if l==0 :
            l = 1

        a = Res.split(" ")  


        ResReal = []
        newstr = ""
        i = 0
        for elem in a :
            ResReal.append([])
            for symb in elem :
                if symb != "a" :
                    newstr = newstr + symb
                else :
                    if newstr != "" :
                        ResReal[i].append(int(newstr))
                        newstr = ""
            if newstr != "" :
                ResReal[i].append(int(newstr))
                newstr = ""
            i += 1



        i = 0
        ResultatTuTBudet = []
        for elem in ResReal :
            ResultatTuTBudet.append([])
            for number in elem :
                ResultatTuTBudet[i].append(T[number-1][0])
            i += 1
        return ResultatTuTBudet 
    data = MaskMethod(T)
    json_data=json.dumps(data, cls=DjangoJSONEncoder)
    return HttpResponse(json_data, content_type='application/json')